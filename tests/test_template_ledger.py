"""Audit ledger template detection and rendering."""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from union_ledger.models.entities import Evidence
from union_ledger.models.enums import EvidenceStatus, EvidenceType
from union_ledger.services.artifact import render_settlement_excel
from union_ledger.services.template_ledger import (
    LAYOUT_AUDIT_LEDGER,
    _parse_month_sections,
    detect_audit_ledger_mapping,
)
from union_ledger.services.template_mapping import detect_mapping_schema_from_bytes

FIXTURE = (
    Path(__file__).resolve().parent
    / "fixtures"
    / "settlement_templates"
    / "audit_ledger_sample.xlsx"
)


def test_detect_audit_ledger_mapping_from_sample() -> None:
    file_bytes = FIXTURE.read_bytes()
    mapping = detect_mapping_schema_from_bytes(file_bytes)

    assert mapping["_layout"] == LAYOUT_AUDIT_LEDGER
    assert mapping["A3"] == "title"
    assert mapping["G1"] == "generated_at"
    assert "columns" in mapping["_ledger"]
    assert "B3" not in mapping


def test_parse_month_sections_from_sample() -> None:
    from union_ledger.services.bank_statement import read_excel_sheet_rows

    rows = read_excel_sheet_rows(FIXTURE.read_bytes())
    sections = _parse_month_sections(rows)

    assert sections
    assert sections[0]["month"] == 12
    assert sections[0]["header_row"] < sections[0]["settlement_row"]


def test_render_audit_ledger_preserves_layout_and_fills_december() -> None:
    file_bytes = FIXTURE.read_bytes()
    mapping = detect_audit_ledger_mapping(file_bytes)
    assert mapping is not None

    evidence = Evidence(
        settlement_id=uuid.uuid4(),
        organization_id=uuid.uuid4(),
        evidence_type=EvidenceType.PHYSICAL_RECEIPT,
        status=EvidenceStatus.CONFIRMED,
        source_file_name="receipt.png",
        source_file_path="/tmp/receipt.png",
        extracted_payload={},
        evidence_date=date(2025, 12, 23),
        merchant_name="학생회\n비품구입",
        amount=Decimal("85000"),
        budget_category="비품",
        is_refund=False,
    )

    output = render_settlement_excel(
        template_path=FIXTURE,
        mapping_schema=mapping,
        spec_fields={
            "title": "컴퓨터공학부 학생회 2학기 결산안",
            "generated_at": "2026-06-10T12:00:00+00:00",
        },
        evidences=[evidence],
    )

    wb = load_workbook(io.BytesIO(output))
    sheet = wb.active
    assert sheet.title == "결산안"
    assert sheet["A3"].value == "컴퓨터공학부 학생회 2학기 결산안"
    assert "2026년" in str(sheet["G1"].value)
    assert sheet["A5"].value == "12월 결산"
    assert sheet["A6"].value == "날짜"

    # 구분(col 2) = budget category (no ledger group assigned → flat mode),
    # 항목(col 3) = OCR merchant name — matching real audit ledgers.
    found = False
    for row in range(7, 20):
        if sheet.cell(row=row, column=3).value == "학생회\n비품구입":
            assert sheet.cell(row=row, column=2).value == "비품"
            assert sheet.cell(row=row, column=5).value == 85000.0
            assert sheet.cell(row=row, column=8).value == "12*1"
            found = True
            break
    assert found, "expected December evidence row in audit ledger block"
    wb.close()


def _evidence(
    *,
    merchant: str,
    amount: str,
    day: int,
    group: str | None = None,
    category: str | None = None,
    is_refund: bool = False,
) -> Evidence:
    return Evidence(
        settlement_id=uuid.uuid4(),
        organization_id=uuid.uuid4(),
        evidence_type=EvidenceType.PHYSICAL_RECEIPT,
        status=EvidenceStatus.CONFIRMED,
        source_file_name="receipt.png",
        source_file_path="/tmp/receipt.png",
        extracted_payload={},
        evidence_date=date(2025, 12, day),
        merchant_name=merchant,
        amount=Decimal(amount),
        budget_category=category,
        group_name=group,
        is_refund=is_refund,
    )


def test_render_audit_ledger_groups_by_구분_with_subtotals() -> None:
    file_bytes = FIXTURE.read_bytes()
    mapping = detect_audit_ledger_mapping(file_bytes)
    assert mapping is not None

    evidences = [
        _evidence(merchant="맘스터치", amount="820000", day=5, group="중간고사 간식행사"),
        _evidence(merchant="서울김밥", amount="400000", day=6, group="중간고사 간식행사"),
        _evidence(merchant="다이소", amount="10000", day=7, category="사무용품/비품"),
    ]

    output = render_settlement_excel(
        template_path=FIXTURE,
        mapping_schema=mapping,
        spec_fields={"title": "결산안", "generated_at": "2026-06-10T12:00:00+00:00"},
        evidences=evidences,
    )

    wb = load_workbook(io.BytesIO(output))
    sheet = wb.active

    # Block layout: items carry 구분 on the first row only; each 구분 ends
    # with a '{구분} 소계' row; ungrouped items fall under 미분류.
    assert sheet.cell(row=7, column=2).value == "중간고사 간식행사"
    assert sheet.cell(row=7, column=3).value == "맘스터치"
    # openpyxl persists empty strings as None.
    assert sheet.cell(row=8, column=2).value in (None, "")
    assert sheet.cell(row=8, column=3).value == "서울김밥"
    assert sheet.cell(row=9, column=3).value == "중간고사 간식행사 소계"
    assert sheet.cell(row=9, column=5).value == 1220000.0
    assert sheet.cell(row=10, column=2).value == "미분류"
    assert sheet.cell(row=10, column=3).value == "다이소"
    assert sheet.cell(row=11, column=3).value == "미분류 소계"
    assert sheet.cell(row=11, column=5).value == 10000.0

    # Receipt numbering counts items only (subtotal rows are not numbered).
    assert sheet.cell(row=7, column=8).value == "12*1"
    assert sheet.cell(row=10, column=8).value == "12*3"

    # 'N월 정산' row carries the month totals (formulas don't survive the
    # row delete/insert, so they are written explicitly).
    settlement_row = None
    for row in range(7, 40):
        text = str(sheet.cell(row=row, column=1).value or "")
        if "정산" in text and "12" in text:
            settlement_row = row
            break
    assert settlement_row is not None, "expected 12월 정산 row"
    assert sheet.cell(row=settlement_row, column=5).value == 1230000.0
    wb.close()
