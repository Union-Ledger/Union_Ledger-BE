"""Strict-OOXML template handling for Excel artifact generation.

openpyxl can't read 'Strict Open XML(엄격)' .xlsx files (it sees zero sheets).
render_settlement_excel converts them to Transitional first. We simulate a
strict file by reverse-mapping a normal workbook's namespaces.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

from union_ledger.services.artifact import (
    _STRICT_NAMESPACE_REPLACEMENTS,
    render_settlement_excel,
)


def _make_strict_xlsx(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    wb.active["A1"] = "고정 라벨"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    out = io.BytesIO()
    with (
        zipfile.ZipFile(buf) as zin,
        zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout,
    ):
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith((".xml", ".rels")):
                text = data.decode("utf-8")
                # Reverse the fix: transitional → strict namespaces.
                for strict_uri, transitional_uri in _STRICT_NAMESPACE_REPLACEMENTS:
                    text = text.replace(transitional_uri, strict_uri)
                data = text.encode("utf-8")
            zout.writestr(item, data)

    path = tmp_path / "strict_template.xlsx"
    path.write_bytes(out.getvalue())
    return path


def test_openpyxl_reads_strict_as_zero_sheets(tmp_path: Path) -> None:
    path = _make_strict_xlsx(tmp_path)
    wb = load_workbook(path)
    assert wb.sheetnames == []  # reproduces the production failure


def test_render_settlement_excel_handles_strict_template(tmp_path: Path) -> None:
    path = _make_strict_xlsx(tmp_path)
    data = render_settlement_excel(
        template_path=path,
        mapping_schema={"B2": "title"},
        spec_fields={"title": "2024 결산안"},
    )
    result = load_workbook(io.BytesIO(data))
    assert result.active is not None
    assert result.active["B2"].value == "2024 결산안"
