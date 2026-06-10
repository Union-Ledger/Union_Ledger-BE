"""Settlement template endpoint tests (spec ④ Step 1).

Covers upload/list/get/patch/delete. We use a tiny non-empty payload as a
stand-in for an .xlsx — the endpoint only validates extension + non-empty,
not Excel structure (parser is deferred to a later slice).
"""

from __future__ import annotations

import io

from httpx import AsyncClient

from conftest import auth_headers, create_org_as_admin, signup

# Tiny non-empty payload — content is never parsed by the endpoint, only
# size and extension are checked.
DUMMY_XLSX = b"PK\x03\x04dummy-xlsx-bytes-not-real"


async def _upload_template(
    client: AsyncClient,
    headers: dict[str, str],
    *,
    org_id: str,
    name: str = "기본 템플릿",
    filename: str = "template.xlsx",
    content: bytes = DUMMY_XLSX,
    mapping_schema: str | None = None,
) -> dict:
    files = {"file": (filename, io.BytesIO(content), "application/vnd.ms-excel")}
    data = {"name": name}
    if mapping_schema is not None:
        data["mapping_schema"] = mapping_schema
    resp = await client.post(
        f"/api/v1/organizations/{org_id}/templates",
        headers=headers,
        files=files,
        data=data,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


# --- Upload ---------------------------------------------------------------


async def test_upload_template_requires_auth(client: AsyncClient) -> None:
    files = {"file": ("template.xlsx", io.BytesIO(DUMMY_XLSX))}
    resp = await client.post(
        "/api/v1/organizations/00000000-0000-0000-0000-000000000000/templates",
        files=files,
        data={"name": "x"},
    )
    assert resp.status_code == 401, resp.text


async def test_upload_template_as_admin(client: AsyncClient) -> None:
    await signup(client, email="t_admin@konkuk.ac.kr")
    headers = await auth_headers(client, "t_admin@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)

    template = await _upload_template(
        client,
        headers,
        org_id=org["id"],
        mapping_schema='{"A1": "title", "B1": "amount"}',
    )
    assert template["name"] == "기본 템플릿"
    assert template["original_filename"] == "template.xlsx"
    assert template["is_active"] is True
    assert template["mapping_schema"] == {"A1": "title", "B1": "amount"}


async def test_upload_template_default_empty_mapping(client: AsyncClient) -> None:
    await signup(client, email="t_empty@konkuk.ac.kr")
    headers = await auth_headers(client, "t_empty@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)

    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "제목"
    ws["A2"] = "학년도"
    ws["A3"] = "총 지출액"
    buf = BytesIO()
    wb.save(buf)
    template_bytes = buf.getvalue()

    files = {"file": ("template.xlsx", BytesIO(template_bytes), "application/vnd.ms-excel")}
    resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data={"name": "자동매핑"},
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["mapping_schema"]["B1"] == "title"
    assert body["mapping_schema"]["B2"] == "academic_year"
    assert body["mapping_schema"]["B3"] == "total_evidence_amount"


async def test_upload_template_rejects_non_excel(client: AsyncClient) -> None:
    await signup(client, email="t_pdf@konkuk.ac.kr")
    headers = await auth_headers(client, "t_pdf@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    files = {"file": ("template.pdf", io.BytesIO(b"%PDF-1.4 not really"))}
    resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data={"name": "x"},
    )
    assert resp.status_code == 400, resp.text


async def test_upload_template_rejects_empty_file(client: AsyncClient) -> None:
    await signup(client, email="t_empty_file@konkuk.ac.kr")
    headers = await auth_headers(client, "t_empty_file@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    files = {"file": ("template.xlsx", io.BytesIO(b""))}
    resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data={"name": "x"},
    )
    assert resp.status_code == 400, resp.text


async def test_upload_template_rejects_invalid_mapping_json(client: AsyncClient) -> None:
    await signup(client, email="t_badjson@konkuk.ac.kr")
    headers = await auth_headers(client, "t_badjson@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    files = {"file": ("template.xlsx", io.BytesIO(DUMMY_XLSX))}
    resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data={"name": "x", "mapping_schema": "{not valid json"},
    )
    assert resp.status_code == 400, resp.text


async def test_upload_template_rejects_non_object_mapping(client: AsyncClient) -> None:
    await signup(client, email="t_listmap@konkuk.ac.kr")
    headers = await auth_headers(client, "t_listmap@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    files = {"file": ("template.xlsx", io.BytesIO(DUMMY_XLSX))}
    resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data={"name": "x", "mapping_schema": "[1, 2, 3]"},
    )
    assert resp.status_code == 400, resp.text


# --- List -----------------------------------------------------------------


async def test_list_templates_default_active_only(client: AsyncClient) -> None:
    await signup(client, email="t_list@konkuk.ac.kr")
    headers = await auth_headers(client, "t_list@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    active = await _upload_template(client, headers, org_id=org["id"], name="활성")
    inactive = await _upload_template(client, headers, org_id=org["id"], name="비활성")

    # Deactivate the second one.
    await client.delete(f"/api/v1/templates/{inactive['id']}", headers=headers)

    listing = await client.get(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
    )
    assert listing.status_code == 200
    ids = [t["id"] for t in listing.json()]
    assert active["id"] in ids
    assert inactive["id"] not in ids


async def test_list_templates_include_inactive(client: AsyncClient) -> None:
    await signup(client, email="t_list_all@konkuk.ac.kr")
    headers = await auth_headers(client, "t_list_all@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    active = await _upload_template(client, headers, org_id=org["id"], name="활성")
    inactive = await _upload_template(client, headers, org_id=org["id"], name="비활성")
    await client.delete(f"/api/v1/templates/{inactive['id']}", headers=headers)

    listing = await client.get(
        f"/api/v1/organizations/{org['id']}/templates?include_inactive=true",
        headers=headers,
    )
    assert listing.status_code == 200
    ids = {t["id"] for t in listing.json()}
    assert {active["id"], inactive["id"]}.issubset(ids)


async def test_list_templates_rejects_non_member(client: AsyncClient) -> None:
    await signup(client, email="t_list_owner@konkuk.ac.kr")
    headers_o = await auth_headers(client, "t_list_owner@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers_o)

    await signup(client, email="t_list_outsider@konkuk.ac.kr")
    headers_x = await auth_headers(client, "t_list_outsider@konkuk.ac.kr")
    resp = await client.get(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers_x,
    )
    assert resp.status_code == 403


# --- Detail / Patch / Delete ---------------------------------------------


async def test_get_template_member_only(client: AsyncClient) -> None:
    await signup(client, email="t_get@konkuk.ac.kr")
    headers = await auth_headers(client, "t_get@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    t = await _upload_template(client, headers, org_id=org["id"])

    ok = await client.get(f"/api/v1/templates/{t['id']}", headers=headers)
    assert ok.status_code == 200

    await signup(client, email="t_get_outsider@konkuk.ac.kr")
    headers_x = await auth_headers(client, "t_get_outsider@konkuk.ac.kr")
    forbid = await client.get(f"/api/v1/templates/{t['id']}", headers=headers_x)
    assert forbid.status_code == 403


async def test_get_template_404(client: AsyncClient) -> None:
    await signup(client, email="t_get_404@konkuk.ac.kr")
    headers = await auth_headers(client, "t_get_404@konkuk.ac.kr")
    resp = await client.get(
        "/api/v1/templates/00000000-0000-0000-0000-000000000000",
        headers=headers,
    )
    assert resp.status_code == 404


async def test_patch_template_renames_and_remaps(client: AsyncClient) -> None:
    await signup(client, email="t_patch@konkuk.ac.kr")
    headers = await auth_headers(client, "t_patch@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    t = await _upload_template(client, headers, org_id=org["id"], name="원래 이름")

    resp = await client.patch(
        f"/api/v1/templates/{t['id']}",
        headers=headers,
        json={"name": "새 이름", "mapping_schema": {"C3": "subtotal"}},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["name"] == "새 이름"
    assert body["mapping_schema"] == {"C3": "subtotal"}


async def test_patch_template_rejects_non_treasurer(client: AsyncClient) -> None:
    await signup(client, email="t_p_owner@konkuk.ac.kr")
    headers_o = await auth_headers(client, "t_p_owner@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers_o)
    t = await _upload_template(client, headers_o, org_id=org["id"])

    await signup(client, email="t_p_outsider@konkuk.ac.kr")
    headers_x = await auth_headers(client, "t_p_outsider@konkuk.ac.kr")
    resp = await client.patch(
        f"/api/v1/templates/{t['id']}",
        headers=headers_x,
        json={"name": "hacked"},
    )
    assert resp.status_code == 403


async def test_delete_template_deactivates(client: AsyncClient) -> None:
    await signup(client, email="t_del@konkuk.ac.kr")
    headers = await auth_headers(client, "t_del@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    t = await _upload_template(client, headers, org_id=org["id"])

    resp = await client.delete(f"/api/v1/templates/{t['id']}", headers=headers)
    assert resp.status_code == 200
    assert resp.json()["is_active"] is False


async def test_upload_audit_ledger_ignores_stale_summary_mapping(
    client: AsyncClient,
) -> None:
    import json
    from pathlib import Path

    fixture = (
        Path(__file__).resolve().parent
        / "fixtures"
        / "settlement_templates"
        / "audit_ledger_sample.xlsx"
    )
    if not fixture.is_file():
        return

    await signup(client, email="t_audit@konkuk.ac.kr")
    headers = await auth_headers(client, "t_audit@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)

    files = {
        "file": (
            "audit_ledger_sample.xlsx",
            io.BytesIO(fixture.read_bytes()),
            "application/vnd.ms-excel",
        )
    }
    data = {
        "name": "공과대 예결산안",
        "mapping_schema": json.dumps({"B1": "title", "B2": "academic_year"}),
    }
    resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data=data,
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["mapping_schema"]["_layout"] == "audit_ledger"
    assert body["mapping_schema"]["A3"] == "title"
    assert "B1" not in body["mapping_schema"]


async def test_upload_deactivates_previous_active_template(
    client: AsyncClient,
) -> None:
    await signup(client, email="t_swap@konkuk.ac.kr")
    headers = await auth_headers(client, "t_swap@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)

    first = await _upload_template(client, headers, org_id=org["id"], name="첫 번째")
    second = await _upload_template(client, headers, org_id=org["id"], name="두 번째")

    listing = await client.get(
        f"/api/v1/organizations/{org['id']}/templates?include_inactive=true",
        headers=headers,
    )
    assert listing.status_code == 200, listing.text
    rows = {row["id"]: row for row in listing.json()}
    assert rows[first["id"]]["is_active"] is False
    assert rows[second["id"]]["is_active"] is True
