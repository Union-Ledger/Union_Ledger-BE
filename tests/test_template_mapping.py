"""Tests for settlement template mapping auto-detection."""

from __future__ import annotations

import io

from openpyxl import Workbook

from union_ledger.services.template_mapping import detect_mapping_schema_from_bytes


def _workbook_bytes(rows: list[tuple[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row_idx, (label, _value) in enumerate(rows, start=1):
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = ""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_detect_mapping_from_label_column() -> None:
    file_bytes = _workbook_bytes(
        [
            ("제목", ""),
            ("학년도", ""),
            ("학기", ""),
            ("총 지출액", ""),
            ("증빙 건수", ""),
        ]
    )

    mapping = detect_mapping_schema_from_bytes(file_bytes)

    assert mapping["B1"] == "title"
    assert mapping["B2"] == "academic_year"
    assert mapping["B3"] == "semester"
    assert mapping["B4"] == "total_evidence_amount"
    assert mapping["B5"] == "evidence_count"


def test_detect_mapping_returns_empty_for_unrecognized_layout() -> None:
    file_bytes = _workbook_bytes([("날짜", ""), ("적요", ""), ("출금", "")])

    assert detect_mapping_schema_from_bytes(file_bytes) == {}
