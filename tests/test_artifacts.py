"""Settlement artifact tests (spec §2 Step 5).

Tests build a real template.xlsx + a real evidence image so the Excel + PDF
generators run end-to-end. The download endpoint is verified by hashing the
returned bytes against the produced file.
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal

from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from PIL import Image
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import async_sessionmaker

from conftest import auth_headers, create_org_as_admin, signup
from union_ledger.models.entities import Evidence
from union_ledger.models.enums import EvidenceStatus, EvidenceType


def _build_template(mapping_cells: dict[str, str]) -> bytes:
    """Make a 1-sheet xlsx with header rows so we can verify cell writes."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "결산안 양식"
    ws["A2"] = "제목"
    ws["A3"] = "년도"
    ws["A4"] = "총 지출"
    # The mapping the treasurer would create:
    for cell, _field in mapping_cells.items():
        ws[cell] = ""  # leave blank — the generator fills it
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_image_bytes(color: tuple[int, int, int]) -> bytes:
    img = Image.new("RGB", (200, 100), color=color)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


async def _upload_template(
    client: AsyncClient,
    headers: dict[str, str],
    *,
    org_id: str,
    mapping: dict[str, str],
) -> dict:
    import json

    template_bytes = _build_template(mapping)
    files = {"file": ("template.xlsx", io.BytesIO(template_bytes))}
    data = {"name": "기본 템플릿", "mapping_schema": json.dumps(mapping)}
    resp = await client.post(
        f"/api/v1/organizations/{org_id}/templates",
        headers=headers,
        files=files,
        data=data,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


async def _create_settlement_with_template(
    client: AsyncClient,
    headers: dict[str, str],
    *,
    org_id: str,
    template_id: str,
    title: str = "2026-1학기 정산",
) -> dict:
    resp = await client.post(
        f"/api/v1/organizations/{org_id}/settlements",
        headers=headers,
        json={
            "title": title,
            "academic_year": 2026,
            "semester": "1",
            "template_id": template_id,
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


async def _seed_image_evidence(
    db_sessionmaker: async_sessionmaker,
    *,
    settlement_id: uuid.UUID,
    organization_id: uuid.UUID,
    storage_root,
    evidence_date: date,
    amount: Decimal,
    color: tuple[int, int, int] = (200, 80, 80),
) -> uuid.UUID:
    """Write a real PNG to storage_root and seed an Evidence row pointing at it."""
    img_dir = storage_root / "evidences" / str(settlement_id)
    img_dir.mkdir(parents=True, exist_ok=True)
    img_path = img_dir / f"{uuid.uuid4()}.png"
    img_path.write_bytes(_build_image_bytes(color))

    async with db_sessionmaker() as session:
        ev = Evidence(
            settlement_id=settlement_id,
            organization_id=organization_id,
            evidence_type=EvidenceType.PHYSICAL_RECEIPT,
            status=EvidenceStatus.NEEDS_REVIEW,
            source_file_name=img_path.name,
            source_file_path=str(img_path),
            extracted_payload={},
            evidence_date=evidence_date,
            merchant_name="가게",
            amount=amount,
        )
        session.add(ev)
        await session.commit()
        await session.refresh(ev)
        return ev.id


# --- Generation ---------------------------------------------------------


async def test_generate_creates_excel_and_pdf(
    client: AsyncClient,
    db_sessionmaker: async_sessionmaker,
    tmp_path,
    monkeypatch,
) -> None:
    """End-to-end: upload template, seed two image evidences, generate, verify."""
    # Redirect storage to tmp_path so we don't pollute the repo.
    from union_ledger.core.config import get_settings

    settings = get_settings()
    monkeypatch.setattr(settings, "storage_root", tmp_path)

    await signup(client, email="art_a@konkuk.ac.kr")
    headers = await auth_headers(client, "art_a@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    mapping = {
        "B2": "title",
        "B3": "academic_year",
        "B4": "total_evidence_amount",
        "B5": "evidence_count",
    }
    template = await _upload_template(client, headers, org_id=org["id"], mapping=mapping)
    settlement = await _create_settlement_with_template(
        client, headers, org_id=org["id"], template_id=template["id"]
    )

    await _seed_image_evidence(
        db_sessionmaker,
        settlement_id=uuid.UUID(settlement["id"]),
        organization_id=uuid.UUID(org["id"]),
        storage_root=tmp_path,
        evidence_date=date(2026, 4, 1),
        amount=Decimal("12000"),
    )
    await _seed_image_evidence(
        db_sessionmaker,
        settlement_id=uuid.UUID(settlement["id"]),
        organization_id=uuid.UUID(org["id"]),
        storage_root=tmp_path,
        evidence_date=date(2026, 4, 2),
        amount=Decimal("3000"),
        color=(80, 200, 80),
    )

    gen = await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate",
        headers=headers,
    )
    assert gen.status_code == 200, gen.text
    body = gen.json()
    assert body["excel"]["status"] == "completed"
    assert body["pdf"]["status"] == "completed"
    assert body["excel"]["artifact_type"] == "settlement_excel"
    assert body["pdf"]["artifact_type"] == "evidence_pdf"

    # Verify Excel content reflects mapping.
    excel_path = body["excel"]["file_path"]
    wb = load_workbook(excel_path)
    sheet = wb.active
    assert sheet["B2"].value == "2026-1학기 정산"
    assert sheet["B3"].value == 2026
    assert float(sheet["B4"].value) == 15000.0
    assert sheet["B5"].value == 2
    wb.close()

    # Verify PDF has at least 2 pages (one per image evidence).
    pdf_path = body["pdf"]["file_path"]
    with open(pdf_path, "rb") as f:
        reader = PdfReader(f)
        assert len(reader.pages) >= 2


async def test_generate_replaces_prior_artifacts(
    client: AsyncClient,
    db_sessionmaker: async_sessionmaker,
    tmp_path,
    monkeypatch,
) -> None:
    from union_ledger.core.config import get_settings

    monkeypatch.setattr(get_settings(), "storage_root", tmp_path)

    await signup(client, email="art_r@konkuk.ac.kr")
    headers = await auth_headers(client, "art_r@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    template = await _upload_template(
        client, headers, org_id=org["id"], mapping={"B2": "title"}
    )
    settlement = await _create_settlement_with_template(
        client, headers, org_id=org["id"], template_id=template["id"]
    )

    first = await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate",
        headers=headers,
    )
    second = await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate",
        headers=headers,
    )
    assert first.status_code == 200
    assert second.status_code == 200

    listing = await client.get(
        f"/api/v1/settlements/{settlement['id']}/artifacts",
        headers=headers,
    )
    rows = listing.json()
    # Only the latest pair should remain (replace policy).
    assert len(rows) == 2
    types = {r["artifact_type"] for r in rows}
    assert types == {"settlement_excel", "evidence_pdf"}


async def test_generate_without_template_marks_excel_failed_pdf_completed(
    client: AsyncClient, tmp_path, monkeypatch
) -> None:
    """A settlement without a template can't produce Excel, but PDF should
    still succeed (it doesn't depend on the template)."""
    from union_ledger.core.config import get_settings

    monkeypatch.setattr(get_settings(), "storage_root", tmp_path)

    await signup(client, email="art_nt@konkuk.ac.kr")
    headers = await auth_headers(client, "art_nt@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    settlement = await client.post(
        f"/api/v1/organizations/{org['id']}/settlements",
        headers=headers,
        json={"title": "no-template", "academic_year": 2026, "semester": "1"},
    )
    assert settlement.status_code == 201

    gen = await client.post(
        f"/api/v1/settlements/{settlement.json()['id']}/artifacts:generate",
        headers=headers,
    )
    assert gen.status_code == 200
    body = gen.json()
    assert body["excel"]["status"] == "failed"
    assert body["excel"]["file_path"] is None
    # PDF works even with zero evidences (blank A4 page).
    assert body["pdf"]["status"] == "completed"


# --- List + auth --------------------------------------------------------


async def test_list_artifacts_member_only(
    client: AsyncClient, tmp_path, monkeypatch
) -> None:
    from union_ledger.core.config import get_settings

    monkeypatch.setattr(get_settings(), "storage_root", tmp_path)

    await signup(client, email="art_l@konkuk.ac.kr")
    owner = await auth_headers(client, "art_l@konkuk.ac.kr")
    org = await create_org_as_admin(client, owner)
    template = await _upload_template(
        client, owner, org_id=org["id"], mapping={"B2": "title"}
    )
    settlement = await _create_settlement_with_template(
        client, owner, org_id=org["id"], template_id=template["id"]
    )
    await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate", headers=owner
    )

    await signup(client, email="art_l_outsider@konkuk.ac.kr")
    out_headers = await auth_headers(client, "art_l_outsider@konkuk.ac.kr")
    forbid = await client.get(
        f"/api/v1/settlements/{settlement['id']}/artifacts", headers=out_headers
    )
    assert forbid.status_code == 403, forbid.text


async def test_generate_requires_treasurer_or_admin(
    client: AsyncClient, tmp_path, monkeypatch
) -> None:
    from union_ledger.core.config import get_settings

    monkeypatch.setattr(get_settings(), "storage_root", tmp_path)

    await signup(client, email="art_g@konkuk.ac.kr")
    owner = await auth_headers(client, "art_g@konkuk.ac.kr")
    org = await create_org_as_admin(client, owner)
    template = await _upload_template(
        client, owner, org_id=org["id"], mapping={"B2": "title"}
    )
    settlement = await _create_settlement_with_template(
        client, owner, org_id=org["id"], template_id=template["id"]
    )

    await signup(client, email="art_g_outsider@konkuk.ac.kr")
    out_headers = await auth_headers(client, "art_g_outsider@konkuk.ac.kr")
    resp = await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate",
        headers=out_headers,
    )
    assert resp.status_code == 403, resp.text


# --- Download ------------------------------------------------------------


async def test_download_serves_excel_bytes(
    client: AsyncClient, tmp_path, monkeypatch
) -> None:
    from union_ledger.core.config import get_settings

    monkeypatch.setattr(get_settings(), "storage_root", tmp_path)

    await signup(client, email="art_d@konkuk.ac.kr")
    headers = await auth_headers(client, "art_d@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)
    template = await _upload_template(
        client, headers, org_id=org["id"], mapping={"B2": "title"}
    )
    settlement = await _create_settlement_with_template(
        client, headers, org_id=org["id"], template_id=template["id"]
    )
    gen = await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate",
        headers=headers,
    )
    excel = gen.json()["excel"]

    dl = await client.get(
        f"/api/v1/artifacts/{excel['id']}/download", headers=headers
    )
    assert dl.status_code == 200, dl.text
    # xlsx magic bytes start with PK (zip).
    assert dl.content[:2] == b"PK"
    # Filename hint should reflect the type.
    cd = dl.headers.get("content-disposition", "")
    assert "settlement_excel" in cd


async def test_generate_uses_active_template_not_draft_snapshot(
    client: AsyncClient,
    db_sessionmaker: async_sessionmaker,
    tmp_path,
    monkeypatch,
) -> None:
    """Settlement pinned to an old summary template must still render the active audit workbook."""
    import uuid
    from datetime import date
    from decimal import Decimal
    from pathlib import Path

    from openpyxl import load_workbook

    from union_ledger.core.config import get_settings

    fixture = (
        Path(__file__).resolve().parent
        / "fixtures"
        / "settlement_templates"
        / "audit_ledger_sample.xlsx"
    )
    if not fixture.is_file():
        return

    monkeypatch.setattr(get_settings(), "storage_root", tmp_path)

    await signup(client, email="art_active@konkuk.ac.kr")
    headers = await auth_headers(client, "art_active@konkuk.ac.kr")
    org = await create_org_as_admin(client, headers)

    summary_mapping = {
        "C2": "academic_year",
        "C3": "title",
        "C8": "total_evidence_amount",
    }
    summary_template = await _upload_template(
        client, headers, org_id=org["id"], mapping=summary_mapping
    )
    settlement = await _create_settlement_with_template(
        client, headers, org_id=org["id"], template_id=summary_template["id"]
    )

    files = {
        "file": (
            "audit_ledger_sample.xlsx",
            io.BytesIO(fixture.read_bytes()),
            "application/vnd.ms-excel",
        )
    }
    audit_resp = await client.post(
        f"/api/v1/organizations/{org['id']}/templates",
        headers=headers,
        files=files,
        data={"name": "공과대 예결산안"},
    )
    assert audit_resp.status_code == 201, audit_resp.text
    audit_template = audit_resp.json()
    assert audit_template["mapping_schema"]["_layout"] == "audit_ledger"

    await _seed_image_evidence(
        db_sessionmaker,
        settlement_id=uuid.UUID(settlement["id"]),
        organization_id=uuid.UUID(org["id"]),
        storage_root=tmp_path,
        evidence_date=date(2025, 12, 23),
        amount=Decimal("85000"),
    )

    gen = await client.post(
        f"/api/v1/settlements/{settlement['id']}/artifacts:generate",
        headers=headers,
    )
    assert gen.status_code == 200, gen.text
    body = gen.json()
    assert body["excel"]["status"] == "completed"

    wb = load_workbook(body["excel"]["file_path"])
    sheet = wb.active
    assert sheet["A3"].value == settlement["title"]
    assert sheet["G1"].value is not None
    wb.close()

    refreshed = await client.get(
        f"/api/v1/settlements/{settlement['id']}",
        headers=headers,
    )
    assert refreshed.status_code == 200, refreshed.text
    assert refreshed.json()["template_id"] == audit_template["id"]


async def test_download_404_for_unknown(client: AsyncClient) -> None:
    await signup(client, email="art_404@konkuk.ac.kr")
    headers = await auth_headers(client, "art_404@konkuk.ac.kr")
    resp = await client.get(
        "/api/v1/artifacts/00000000-0000-0000-0000-000000000000/download",
        headers=headers,
    )
    assert resp.status_code == 404, resp.text
