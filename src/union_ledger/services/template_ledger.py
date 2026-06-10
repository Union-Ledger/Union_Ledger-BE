"""Audit-style settlement ledger templates (공과대학 예결산안 등).

These workbooks use repeating monthly blocks:

    N월 결산
    날짜 | 구분 | 항목 | 수입 | 지출 | 잔액 | 비고 | 영수증
    ... data rows ...
    N월 정산

Auto-detection stores ``_layout`` / ``_ledger`` metadata in ``mapping_schema``.
Generation writes header cells and fills each month block with evidences.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from union_ledger.models.entities import Evidence
from union_ledger.services.bank_statement import read_excel_sheet_rows

LAYOUT_AUDIT_LEDGER = "audit_ledger"

_LEDGER_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("날짜",),
    "category": ("구분",),
    "item": ("항목",),
    "income": ("수입",),
    "expense": ("지출",),
    "balance": ("잔액",),
    "memo": ("비고",),
    "receipt": ("영수증",),
}

_SECTION_OPEN = re.compile(r"^(\d+)\s*월\s*결산")
_SECTION_CLOSE = re.compile(r"^(\d+)\s*월\s*정산")


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "").lower()


def _find_ledger_header_row(rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
    required = ("date", "category", "item", "income", "expense")
    for row_idx, row in enumerate(rows[:120]):
        columns: dict[str, int] = {}
        for col_idx, raw in enumerate(row):
            cell = _normalize_header(raw)
            if not cell:
                continue
            for key, aliases in _LEDGER_HEADER_ALIASES.items():
                if key in columns:
                    continue
                if any(alias in cell for alias in aliases):
                    columns[key] = col_idx
        if all(key in columns for key in required):
            return row_idx, columns
    return None


def detect_audit_ledger_mapping(file_bytes: bytes) -> dict[str, Any] | None:
    """Return mapping metadata when the workbook looks like an audit ledger."""
    try:
        rows = read_excel_sheet_rows(file_bytes)
    except Exception:
        return None

    header = _find_ledger_header_row(rows)
    if header is None:
        return None

    _, columns = header
    return {
        "_layout": LAYOUT_AUDIT_LEDGER,
        "A3": "title",
        "G1": "generated_at",
        "_ledger": {"columns": columns},
    }


def format_korean_header_date(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
    elif isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value
    else:
        return str(value)
    return f"{dt.year}년 {dt.month}월 {dt.day}일"


def format_ledger_date(value: date | None) -> str:
    if value is None:
        return ""
    return f"{value.year % 100:02d}. {value.month:02d}. {value.day:02d}."


def _cell_text(row: list[object], col_idx: int | None) -> str:
    if col_idx is None or col_idx >= len(row):
        return ""
    raw = row[col_idx]
    if raw is None:
        return ""
    return str(raw).strip()


def _parse_month_sections(rows: list[list[object]]) -> list[dict[str, int]]:
    sections: list[dict[str, int]] = []
    idx = 0
    while idx < len(rows):
        title = _cell_text(rows[idx], 0)
        open_match = _SECTION_OPEN.match(title)
        if not open_match:
            idx += 1
            continue
        month = int(open_match.group(1))
        header_row: int | None = None
        settlement_row: int | None = None
        scan_end = min(len(rows), idx + 8)
        for r in range(idx + 1, scan_end):
            row = rows[r]
            header_probe = _find_ledger_header_row([row])
            if header_probe is not None and header_probe[0] == 0:
                header_row = r
                break
        if header_row is None:
            idx += 1
            continue
        for r in range(header_row + 1, min(len(rows), header_row + 40)):
            close_match = _SECTION_CLOSE.match(_cell_text(rows[r], 0))
            if close_match and int(close_match.group(1)) == month:
                settlement_row = r
                break
        if settlement_row is not None:
            sections.append(
                {
                    "month": month,
                    "title_row": idx,
                    "header_row": header_row,
                    "settlement_row": settlement_row,
                }
            )
            idx = settlement_row + 1
            continue
        idx += 1
    return sections


def _write_cell(sheet: Worksheet, row: int, col_idx: int, value: object) -> None:
    sheet.cell(row=row, column=col_idx + 1, value=value)


UNGROUPED_LABEL = "미분류"


def _group_key(evidence: Evidence) -> str:
    """구분 grouping key — empty string when the treasurer didn't assign one."""
    return (evidence.group_name or "").strip()


def _evidence_income_expense(evidence: Evidence) -> tuple[Decimal, Decimal]:
    magnitude = abs(evidence.amount or Decimal(0))
    if evidence.is_refund:
        return magnitude, Decimal(0)
    return Decimal(0), magnitude


def _write_evidence_row(
    sheet: Worksheet,
    *,
    row: int,
    evidence: Evidence,
    columns: dict[str, int],
    seq: int,
    month: int,
    running_balance: Decimal,
    category_label: str | None,
) -> Decimal:
    """Write one ledger item row.

    구분(category) column carries the ledger group (행사/용도) — passed in by the
    caller so it appears only on the first row of a group, like the sample
    ledgers. 항목(item) column is the merchant from OCR.
    """
    income, expense = _evidence_income_expense(evidence)
    running_balance = running_balance + income - expense

    if "date" in columns:
        _write_cell(sheet, row, columns["date"], format_ledger_date(evidence.evidence_date))
    if "category" in columns:
        _write_cell(sheet, row, columns["category"], category_label or "")
    if "item" in columns:
        _write_cell(
            sheet,
            row,
            columns["item"],
            evidence.merchant_name or evidence.budget_category or "",
        )
    if "income" in columns:
        _write_cell(sheet, row, columns["income"], float(income) if income else None)
    if "expense" in columns:
        _write_cell(sheet, row, columns["expense"], float(expense) if expense else None)
    if "balance" in columns:
        _write_cell(sheet, row, columns["balance"], float(running_balance))
    if "memo" in columns:
        _write_cell(sheet, row, columns["memo"], "")
    if "receipt" in columns:
        _write_cell(sheet, row, columns["receipt"], f"{month}*{seq}")

    return running_balance


def _write_subtotal_row(
    sheet: Worksheet,
    *,
    row: int,
    columns: dict[str, int],
    label: str,
    income: Decimal,
    expense: Decimal,
) -> None:
    """구분 소계 row — '{구분} 소계' with the group's income/expense totals."""
    if "item" in columns:
        _write_cell(sheet, row, columns["item"], f"{label} 소계")
    if "income" in columns:
        _write_cell(sheet, row, columns["income"], float(income))
    if "expense" in columns:
        _write_cell(sheet, row, columns["expense"], float(expense))


def fill_audit_ledger(
    sheet: Worksheet,
    *,
    evidences: list[Evidence],
    ledger_config: dict[str, Any],
) -> None:
    """Replace example ledger rows with settlement evidences, grouped by month."""
    columns = ledger_config.get("columns")
    if not isinstance(columns, dict):
        return

    normalized_columns = {str(k): int(v) for k, v in columns.items()}

    max_row = sheet.max_row or 1
    rows: list[list[object]] = []
    for row in sheet.iter_rows(min_row=1, max_row=max_row, values_only=True):
        rows.append(list(row))

    sections = _parse_month_sections(rows)
    if not sections:
        return

    by_month: dict[int, list[Evidence]] = {}
    unassigned: list[Evidence] = []
    for evidence in sorted(
        evidences,
        key=lambda ev: (
            ev.evidence_date or date.min,
            ev.created_at,
        ),
    ):
        if evidence.evidence_date is None:
            unassigned.append(evidence)
            continue
        by_month.setdefault(evidence.evidence_date.month, []).append(evidence)
    if unassigned and sections:
        by_month.setdefault(sections[0]["month"], []).extend(unassigned)

    for section in sorted(sections, key=lambda s: s["settlement_row"], reverse=True):
        month = section["month"]
        month_evidences = by_month.get(month, [])
        if not month_evidences:
            continue

        header_row = section["header_row"] + 1
        settlement_row = section["settlement_row"] + 1

        # 구분(group) blocks in first-appearance order (evidences are already
        # date-sorted). Items without a 구분 fall under 미분류 — but only when
        # the month actually uses groups; otherwise we keep flat rows with the
        # budget category in the 구분 column (legacy data stays sensible).
        groups: list[tuple[str, list[Evidence]]] = []
        group_index: dict[str, int] = {}
        for evidence in month_evidences:
            key = _group_key(evidence)
            if key not in group_index:
                group_index[key] = len(groups)
                groups.append((key, []))
            groups[group_index[key]][1].append(evidence)
        emit_subtotals = any(key for key, _ in groups)

        total_rows = len(month_evidences) + (len(groups) if emit_subtotals else 0)

        delete_start = header_row + 1
        delete_count = settlement_row - delete_start

        # The template's example rows often merge the 구분 column across a
        # block. delete_rows/insert_rows don't reliably adjust merged ranges,
        # leaving phantom merges over our data rows — unmerge anything that
        # intersects the example region before mutating it.
        for merged_range in list(sheet.merged_cells.ranges):
            if (
                merged_range.max_row >= delete_start
                and merged_range.min_row <= settlement_row - 1
            ):
                sheet.unmerge_cells(str(merged_range))

        if delete_count > 0:
            sheet.delete_rows(delete_start, delete_count)

        insert_at = header_row + 1
        sheet.insert_rows(insert_at, total_rows)

        running_balance = Decimal(0)
        month_income = Decimal(0)
        month_expense = Decimal(0)
        row = insert_at
        seq = 0
        for key, items in groups:
            group_label = key or UNGROUPED_LABEL
            group_income = Decimal(0)
            group_expense = Decimal(0)
            for position, evidence in enumerate(items):
                seq += 1
                if emit_subtotals:
                    # 구분 appears on the first row of its block (sample style).
                    category_label = group_label if position == 0 else ""
                else:
                    category_label = evidence.budget_category or ""
                income, expense = _evidence_income_expense(evidence)
                group_income += income
                group_expense += expense
                running_balance = _write_evidence_row(
                    sheet,
                    row=row,
                    evidence=evidence,
                    columns=normalized_columns,
                    seq=seq,
                    month=month,
                    running_balance=running_balance,
                    category_label=category_label,
                )
                row += 1
            month_income += group_income
            month_expense += group_expense
            if emit_subtotals:
                _write_subtotal_row(
                    sheet,
                    row=row,
                    columns=normalized_columns,
                    label=group_label,
                    income=group_income,
                    expense=group_expense,
                )
                row += 1

        # 'N월 정산' row — write the month totals explicitly (template formulas
        # don't survive the row delete/insert above).
        new_settlement_row = settlement_row - delete_count + total_rows
        if "income" in normalized_columns:
            _write_cell(
                sheet, new_settlement_row, normalized_columns["income"], float(month_income)
            )
        if "expense" in normalized_columns:
            _write_cell(
                sheet, new_settlement_row, normalized_columns["expense"], float(month_expense)
            )
        if "balance" in normalized_columns:
            _write_cell(
                sheet, new_settlement_row, normalized_columns["balance"], float(running_balance)
            )
