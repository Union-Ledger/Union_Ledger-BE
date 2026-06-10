"""Settlement artifact generation (spec §2 Step 5).

Two outputs per settlement:
  1. SETTLEMENT_EXCEL — clones the org's template.xlsx and writes computed
     values into cells declared in `template.mapping_schema`. The mapping is
     `{cell_address: spec_field}` (e.g. `{"B2": "title", "D5": "total_amount"}`).
  2. EVIDENCE_PDF — concatenates every evidence file in date order. Image
     evidences (jpg/png/...) are inserted as one image per page; PDF evidences
     (거래명세서/전자영수증) are merged page-by-page via pypdf.

Status transitions per artifact row:
    QUEUED → PROCESSING → COMPLETED
                       ↘ FAILED (with error message in file_path? — no, we
                                  leave file_path NULL on failure)

Generation runs synchronously in a thread (`asyncio.to_thread`). The MVP
spec doesn't mandate a queue; if a settlement has thousands of evidences
this would block, but for student-org scale it's fine.

Re-run policy:
  POST :generate first deletes all prior artifact rows for the settlement,
  then creates a fresh pair. The on-disk file from the prior run is left in
  place — cheap and avoids racing a concurrent download.
"""

from __future__ import annotations

import io
import uuid
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.exceptions import CellCoordinatesException, InvalidFileException
from PIL import Image, UnidentifiedImageError
from pypdf import PdfReader, PdfWriter
from pypdf.errors import PdfReadError
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from union_ledger.core.config import Settings
from union_ledger.models.entities import (
    BankStatementUpload,
    BankTransaction,
    Evidence,
    Organization,
    ReconciliationResult,
    Settlement,
    SettlementArtifact,
    SettlementTemplate,
)
from union_ledger.models.enums import (
    ArtifactStatus,
    ArtifactType,
    MatchStatus,
)
from union_ledger.services.template_ledger import (
    LAYOUT_AUDIT_LEDGER,
    fill_audit_ledger,
    format_korean_header_date,
)
from union_ledger.services.settlement_template import (
    TemplateNotFound as SettlementTemplateNotFound,
    effective_mapping_schema,
    resolve_template_for_generation,
)

# Image suffixes we know PIL can convert to PDF without further work.
_IMAGE_SUFFIXES = {".bmp", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
_PDF_SUFFIXES = {".pdf"}

# Spec field names recognized in `mapping_schema`. The treasurer authors the
# template once; their mapping points each spec field to a concrete cell.
_SUPPORTED_SPEC_FIELDS = frozenset(
    {
        "title",
        "academic_year",
        "semester",
        "organization_name",
        "college_name",
        "department_name",
        "total_evidence_amount",
        "evidence_count",
        "bank_transaction_count",
        "matched_count",
        "amount_mismatch_count",
        "date_mismatch_count",
        "missing_bank_transaction_count",
        "missing_evidence_count",
        "submitted_at",
        "audited_at",
        "published_at",
        "generated_at",
    }
)


class ArtifactError(Exception):
    """Base error (maps to HTTP 4xx)."""


class ArtifactNotFound(ArtifactError):
    pass


class ArtifactNotReady(ArtifactError):
    """Tried to download before generation completed."""


class ArtifactFileMissing(ArtifactError):
    """DB row marks COMPLETED but the file is missing on disk."""


class TemplateMissing(ArtifactError):
    """Settlement has no associated template — Excel can't be generated."""


@dataclass(slots=True)
class GenerationOutcome:
    excel: SettlementArtifact
    pdf: SettlementArtifact
    excel_error: str | None = None
    pdf_error: str | None = None


# --- Spec-field aggregation ---------------------------------------------


@dataclass(slots=True)
class _SettlementSnapshot:
    settlement: Settlement
    organization: Organization
    evidences: Sequence[Evidence]
    bank_transactions: Sequence[BankTransaction]
    reconciliation: Sequence[ReconciliationResult]


async def _load_snapshot(
    session: AsyncSession, *, settlement: Settlement
) -> _SettlementSnapshot:
    organization = await session.get(Organization, settlement.organization_id)
    if organization is None:
        raise ArtifactError("결산안에 연결된 조직을 찾을 수 없습니다.")
    evidences = (
        await session.scalars(
            select(Evidence)
            .where(Evidence.settlement_id == settlement.id)
            .order_by(
                Evidence.evidence_date.asc().nullslast(),
                Evidence.created_at.asc(),
            )
        )
    ).all()
    bank_txs = (
        await session.scalars(
            select(BankTransaction)
            .join(
                BankStatementUpload,
                BankTransaction.upload_id == BankStatementUpload.id,
            )
            .where(BankStatementUpload.settlement_id == settlement.id)
        )
    ).all()
    reconciliation = (
        await session.scalars(
            select(ReconciliationResult).where(
                ReconciliationResult.settlement_id == settlement.id
            )
        )
    ).all()
    return _SettlementSnapshot(
        settlement=settlement,
        organization=organization,
        evidences=list(evidences),
        bank_transactions=list(bank_txs),
        reconciliation=list(reconciliation),
    )


def _compute_spec_fields(snapshot: _SettlementSnapshot) -> dict[str, Any]:
    """Compile every supported spec field from the loaded snapshot."""
    s = snapshot.settlement
    org = snapshot.organization

    total_amount = Decimal(0)
    for ev in snapshot.evidences:
        if ev.amount is not None:
            # Refunds (환불) subtract so the artifact total nets out.
            total_amount += -abs(ev.amount) if ev.is_refund else abs(ev.amount)

    rec_counts = {status: 0 for status in MatchStatus}
    for r in snapshot.reconciliation:
        rec_counts[r.status] = rec_counts.get(r.status, 0) + 1

    return {
        "title": s.title,
        "academic_year": s.academic_year,
        "semester": s.semester,
        "organization_name": org.name,
        "college_name": org.college_name,
        "department_name": org.department_name,
        "total_evidence_amount": total_amount,
        "evidence_count": len(snapshot.evidences),
        "bank_transaction_count": len(snapshot.bank_transactions),
        "matched_count": rec_counts.get(MatchStatus.MATCHED, 0),
        "amount_mismatch_count": rec_counts.get(MatchStatus.AMOUNT_MISMATCH, 0),
        "date_mismatch_count": rec_counts.get(MatchStatus.DATE_MISMATCH, 0),
        "missing_bank_transaction_count": rec_counts.get(
            MatchStatus.MISSING_BANK_TRANSACTION, 0
        ),
        "missing_evidence_count": rec_counts.get(MatchStatus.MISSING_EVIDENCE, 0),
        "submitted_at": _iso_or_none(s.submitted_at),
        "audited_at": _iso_or_none(s.audited_at),
        "published_at": _iso_or_none(s.published_at),
        "generated_at": datetime.now(UTC).isoformat(timespec="seconds"),
    }


def _iso_or_none(value: datetime | date | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return value.isoformat()


# --- Excel generation ----------------------------------------------------


def _is_valid_cell_address(addr: str) -> bool:
    """Strict validation — accept only A1-style references (column letters + row)."""
    try:
        col, row = coordinate_from_string(addr)
        column_index_from_string(col)
        return row >= 1
    except (ValueError, TypeError, CellCoordinatesException):
        return False


# 'Strict Open XML(엄격)' uses different namespace URIs than the Transitional
# format openpyxl supports — so openpyxl reads a strict .xlsx as having ZERO
# sheets. Excel (esp. on Mac) can save in this format. We rewrite the strict
# URIs back to transitional so the template (layout intact) becomes readable.
_STRICT_NAMESPACE_REPLACEMENTS = (
    (
        "purl.oclc.org/ooxml/spreadsheetml/main",
        "schemas.openxmlformats.org/spreadsheetml/2006/main",
    ),
    (
        "purl.oclc.org/ooxml/officeDocument/relationships",
        "schemas.openxmlformats.org/officeDocument/2006/relationships",
    ),
    (
        "purl.oclc.org/ooxml/drawingml/main",
        "schemas.openxmlformats.org/drawingml/2006/main",
    ),
    (
        "purl.oclc.org/ooxml/officeDocument/extendedProperties",
        "schemas.openxmlformats.org/officeDocument/2006/extended-properties",
    ),
    (
        "purl.oclc.org/ooxml/officeDocument/customProperties",
        "schemas.openxmlformats.org/officeDocument/2006/custom-properties",
    ),
    (
        "purl.oclc.org/ooxml/officeDocument/docPropsVTypes",
        "schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes",
    ),
)


def _convert_strict_part(data: bytes) -> bytes:
    text = data.decode("utf-8", "replace")
    for strict_uri, transitional_uri in _STRICT_NAMESPACE_REPLACEMENTS:
        text = text.replace(strict_uri, transitional_uri)
    text = text.replace(' conformance="strict"', "")
    return text.encode("utf-8")


def _convert_strict_xlsx(template_path: Path) -> io.BytesIO:
    """Rewrite a Strict-OOXML .xlsx into the Transitional format openpyxl reads."""
    buf = io.BytesIO()
    with (
        zipfile.ZipFile(template_path) as zin,
        zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout,
    ):
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith((".xml", ".rels")):
                data = _convert_strict_part(data)
            zout.writestr(item, data)
    buf.seek(0)
    return buf


def _load_template_workbook(template_path: Path):
    """Load a template workbook, transparently handling Strict-OOXML files."""
    try:
        wb = load_workbook(template_path)
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
        raise ArtifactError(f"템플릿 엑셀을 열 수 없습니다: {exc}") from exc

    if wb.sheetnames:
        return wb

    # Zero sheets → almost certainly Strict OOXML. Convert and reload.
    wb.close()
    try:
        converted = _convert_strict_xlsx(template_path)
        wb = load_workbook(converted)
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
        raise ArtifactError(
            "템플릿에서 시트를 읽지 못했습니다. 'Strict Open XML(엄격)' 형식일 수 "
            f"있는데 자동 변환에 실패했습니다 ({exc}). Excel에서 'Excel 통합 "
            "문서(.xlsx)'로 다시 저장해 양식을 재업로드해 주세요."
        ) from exc

    if not wb.sheetnames:
        wb.close()
        raise ArtifactError(
            "템플릿에서 시트를 읽지 못했습니다. Excel에서 'Excel 통합 문서(.xlsx)'로 "
            "다시 저장해 양식을 재업로드해 주세요."
        )
    return wb


def render_settlement_excel(
    *,
    template_path: Path,
    mapping_schema: dict[str, Any],
    spec_fields: dict[str, Any],
    evidences: Sequence[Evidence] | None = None,
) -> bytes:
    """Open template, write each `mapping[cell] = spec_field` value, return bytes."""
    wb = _load_template_workbook(template_path)

    sheet = wb.active
    if sheet is None:
        wb.close()
        raise ArtifactError("템플릿에 활성 시트가 없습니다.")

    layout = mapping_schema.get("_layout")
    ledger_config = mapping_schema.get("_ledger")
    evidence_rows = list(evidences or [])

    for cell_addr, field_name in mapping_schema.items():
        if not isinstance(cell_addr, str) or not _is_valid_cell_address(cell_addr):
            continue
        if not isinstance(field_name, str) or field_name not in _SUPPORTED_SPEC_FIELDS:
            continue
        value = spec_fields.get(field_name)
        if layout == LAYOUT_AUDIT_LEDGER and field_name == "generated_at":
            value = format_korean_header_date(value)
        if isinstance(value, Decimal):
            value = float(value)
        sheet[cell_addr] = value

    if (
        layout == LAYOUT_AUDIT_LEDGER
        and isinstance(ledger_config, dict)
        and evidence_rows
    ):
        fill_audit_ledger(
            sheet,
            evidences=list(evidence_rows),
            ledger_config=ledger_config,
        )

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# --- PDF generation ------------------------------------------------------


def render_evidence_pdf(*, evidence_paths: Sequence[Path]) -> bytes:
    """Build a PDF by concatenating every evidence file in the order given.

    For images: each image becomes one page. For PDFs: every page in the
    source PDF is appended as-is. Files we can't read (missing, corrupt,
    unsupported format) are silently skipped so one bad evidence doesn't
    fail the whole batch — the resulting PDF is still useful for the rest.
    """
    writer = PdfWriter()

    for path in evidence_paths:
        if not path.exists():
            continue
        suffix = path.suffix.lower()
        if suffix in _IMAGE_SUFFIXES:
            page_pdf = _image_to_pdf_bytes(path)
            if page_pdf is None:
                continue
            try:
                writer.append(PdfReader(io.BytesIO(page_pdf)))
            except PdfReadError:
                continue
        elif suffix in _PDF_SUFFIXES:
            try:
                writer.append(PdfReader(path))
            except (PdfReadError, OSError):
                continue
        # else: unknown suffix — skip.

    if len(writer.pages) == 0:
        # Even with no evidences we want a valid, empty PDF for the FE to
        # download (so the caller knows generation didn't crash).
        writer.add_blank_page(width=595, height=842)  # A4 in points

    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


def _image_to_pdf_bytes(path: Path) -> bytes | None:
    """Convert a single image file into a 1-page PDF in memory."""
    try:
        with Image.open(path) as img:
            rgb = img.convert("RGB")
            buf = io.BytesIO()
            rgb.save(buf, format="PDF")
            return buf.getvalue()
    except (UnidentifiedImageError, OSError):
        return None


# --- Persistence helpers -------------------------------------------------


def _artifact_dir(settings: Settings, settlement_id: uuid.UUID) -> Path:
    path = (settings.storage_root / "artifacts" / str(settlement_id)).resolve()
    path.mkdir(parents=True, exist_ok=True)
    return path


def _write_bytes(directory: Path, suffix: str, data: bytes) -> Path:
    path = directory / f"{uuid.uuid4()}{suffix}"
    path.write_bytes(data)
    return path


# --- Orchestrator --------------------------------------------------------


async def generate_settlement_artifacts(
    session: AsyncSession,
    *,
    settings: Settings,
    settlement: Settlement,
) -> GenerationOutcome:
    # Replace prior artifacts: the FE always reads the latest pair.
    await session.execute(
        delete(SettlementArtifact).where(
            SettlementArtifact.settlement_id == settlement.id
        )
    )
    excel_row = SettlementArtifact(
        settlement_id=settlement.id,
        artifact_type=ArtifactType.SETTLEMENT_EXCEL,
        status=ArtifactStatus.QUEUED,
    )
    pdf_row = SettlementArtifact(
        settlement_id=settlement.id,
        artifact_type=ArtifactType.EVIDENCE_PDF,
        status=ArtifactStatus.QUEUED,
    )
    session.add(excel_row)
    session.add(pdf_row)
    await session.commit()
    await session.refresh(excel_row)
    await session.refresh(pdf_row)

    snapshot = await _load_snapshot(session, settlement=settlement)
    spec_fields = _compute_spec_fields(snapshot)
    out_dir = _artifact_dir(settings, settlement.id)

    excel_error: str | None = None
    pdf_error: str | None = None

    # Excel
    excel_row.status = ArtifactStatus.PROCESSING
    await session.commit()
    try:
        try:
            template = await resolve_template_for_generation(
                session, settlement=settlement
            )
        except SettlementTemplateNotFound as exc:
            raise TemplateMissing(str(exc)) from exc

        mapping_schema = await effective_mapping_schema(session, template=template)

        excel_bytes = render_settlement_excel(
            template_path=Path(template.file_path),
            mapping_schema=mapping_schema,
            spec_fields=spec_fields,
            evidences=list(snapshot.evidences),
        )
        excel_path = _write_bytes(out_dir, ".xlsx", excel_bytes)
        excel_row.file_path = str(excel_path)
        excel_row.status = ArtifactStatus.COMPLETED
    except (ArtifactError, OSError) as exc:
        # We still attempt the PDF below — the two outputs are independent.
        excel_error = str(exc)
        excel_row.status = ArtifactStatus.FAILED
        excel_row.file_path = None
    await session.commit()

    # PDF
    pdf_row.status = ArtifactStatus.PROCESSING
    await session.commit()
    try:
        evidence_paths = [
            Path(ev.source_file_path) for ev in snapshot.evidences if ev.source_file_path
        ]
        pdf_bytes = render_evidence_pdf(evidence_paths=evidence_paths)
        pdf_path = _write_bytes(out_dir, ".pdf", pdf_bytes)
        pdf_row.file_path = str(pdf_path)
        pdf_row.status = ArtifactStatus.COMPLETED
    except (ArtifactError, OSError) as exc:
        pdf_error = str(exc)
        pdf_row.status = ArtifactStatus.FAILED
        pdf_row.file_path = None
    await session.commit()

    await session.refresh(excel_row)
    await session.refresh(pdf_row)
    return GenerationOutcome(
        excel=excel_row,
        pdf=pdf_row,
        excel_error=excel_error,
        pdf_error=pdf_error,
    )


async def list_settlement_artifacts(
    session: AsyncSession, *, settlement_id: uuid.UUID
) -> Sequence[SettlementArtifact]:
    rows = await session.scalars(
        select(SettlementArtifact)
        .where(SettlementArtifact.settlement_id == settlement_id)
        .order_by(SettlementArtifact.created_at.desc())
    )
    return list(rows.all())


async def get_artifact_or_raise(
    session: AsyncSession, artifact_id: uuid.UUID
) -> SettlementArtifact:
    row = await session.get(SettlementArtifact, artifact_id)
    if row is None:
        raise ArtifactNotFound("산출물을 찾을 수 없습니다.")
    return row


def resolve_download(artifact: SettlementArtifact) -> Path:
    if artifact.status != ArtifactStatus.COMPLETED:
        raise ArtifactNotReady(
            f"산출물이 아직 준비되지 않았습니다 (현재: {artifact.status.value})."
        )
    if artifact.file_path is None:
        raise ArtifactFileMissing("산출물 파일 경로가 비어 있습니다.")
    path = Path(artifact.file_path)
    if not path.exists():
        raise ArtifactFileMissing("산출물 파일이 존재하지 않습니다.")
    return path
