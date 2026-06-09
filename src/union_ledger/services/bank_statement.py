"""Bank statement upload + xlsx parser (spec §2 Step 3).

Korean banks export per-bank custom xlsx layouts (sometimes with a few rows
of metadata before the table header), so the parser walks the sheet to find
the header row by keyword matching, then reads `transaction_date`,
`description`, and `amount` columns.

Kakao Bank email xlsx uses strict OOXML that openpyxl cannot load; we fall back
to reading `xl/worksheets/sheet1.xml` directly. Kakao columns:
  거래일시 | 구분(출금/입금) | 거래금액 | 잔액 | 거래명 | 메모

Heuristic header keywords (case/whitespace insensitive):
  - date         : 거래일시, 거래일, 거래일자, 일자, 날짜, transaction_date, date
  - description  : 거래명, 적요, 내용, 거래내역, description, memo
  - type         : 구분 (출금 → negative 거래금액)
  - withdrawal   : 출금, 출금액, withdrawal
  - deposit      : 입금, 입금액, deposit
  - amount       : 금액, 거래금액, amount

Sign convention for `amount`:
  - If the sheet has a single signed `amount` column, we keep the sign.
  - If it has separate withdrawal/deposit columns, withdrawals become
    negative and deposits stay positive — settlements track expenses, so
    most matched rows will be negative.
"""

from __future__ import annotations

import re
import uuid
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

import xlrd
from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from union_ledger.core.config import Settings
from union_ledger.models.entities import (
    BankStatementUpload,
    BankTransaction,
    ReconciliationResult,
)
from union_ledger.models.enums import BankStatementStatus

# .xls = legacy binary (KB·신한 등 일부 은행이 여전히 .xls로 내려줌), read via xlrd.
SUPPORTED_BANK_STATEMENT_SUFFIXES = {".xlsx", ".xlsm", ".xls"}

# Legacy .xls (BIFF) files are OLE2 compound documents — they start with this
# 8-byte magic. .xlsx/.xlsm are ZIP archives (start with "PK"). We sniff the
# bytes (not the filename) so a mislabeled extension still routes correctly.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_DATE_KEYWORDS = (
    "거래일시",
    "거래일",
    "거래일자",
    "일자",
    "날짜",
    "transaction_date",
    "date",
)
_DESC_KEYWORDS = (
    "거래명",
    "적요",
    "내용",
    "거래내역",
    "description",
    "memo",
    "비고",
)
_MEMO_KEYWORDS = ("메모", "note")
_TYPE_KEYWORDS = ("구분", "입출금", "type")
_WITHDRAWAL_KEYWORDS = ("출금", "출금액", "withdrawal", "출금금액")
_DEPOSIT_KEYWORDS = ("입금", "입금액", "deposit", "입금금액")
_AMOUNT_KEYWORDS = ("금액", "거래금액", "amount")

# Maximum rows to scan when looking for the header row. Korean bank exports
# rarely push the header past row 20.
_MAX_HEADER_SCAN_ROWS = 30


class BankStatementError(Exception):
    """Base for bank-statement-layer errors (maps to HTTP 4xx)."""


class UnsupportedBankStatementFormat(BankStatementError):
    pass


class EmptyBankStatementFile(BankStatementError):
    pass


class BankStatementParseError(BankStatementError):
    pass


class BankStatementUploadNotFound(BankStatementError):
    pass


@dataclass(slots=True)
class StoredBankStatement:
    original_name: str
    absolute_path: Path
    relative_path: Path
    size: int


@dataclass(slots=True)
class ParsedTransaction:
    transaction_date: date
    description: str
    amount: Decimal


def _validate_bank_statement_filename(filename: str) -> str:
    sanitized = Path(filename).name or "bank_statement.xlsx"
    suffix = Path(sanitized).suffix.lower()
    if suffix not in SUPPORTED_BANK_STATEMENT_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_BANK_STATEMENT_SUFFIXES))
        raise UnsupportedBankStatementFormat(
            f"지원하지 않는 거래내역 파일 형식입니다. 지원 확장자: {supported}"
        )
    return sanitized


async def save_bank_statement_file(
    settings: Settings,
    *,
    settlement_id: uuid.UUID,
    upload_file: UploadFile,
) -> StoredBankStatement:
    original_name = _validate_bank_statement_filename(upload_file.filename or "bank_statement.xlsx")
    suffix = Path(original_name).suffix.lower()
    file_bytes = await upload_file.read()
    if not file_bytes:
        raise EmptyBankStatementFile("비어 있는 파일은 업로드할 수 없습니다.")

    relative_path = (
        Path("bank_statements") / str(settlement_id) / f"{uuid.uuid4()}{suffix}"
    )
    absolute_path = (settings.storage_root / relative_path).resolve()
    absolute_path.parent.mkdir(parents=True, exist_ok=True)
    absolute_path.write_bytes(file_bytes)
    return StoredBankStatement(
        original_name=original_name,
        absolute_path=absolute_path,
        relative_path=relative_path,
        size=len(file_bytes),
    )


def _normalize_header_cell(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).lower()


def _header_matches(cell: str, keywords: Iterable[str]) -> bool:
    return any(keyword.lower() in cell for keyword in keywords)


def _coerce_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Common Korean bank exports: "2026-04-29", "2026.04.29", "20260429", and
    # datetime variants like KB's "2026.06.07 16:59:05" (거래일시 = date + time).
    # Kakao Bank xlsx stores Excel strict-ISO datetimes (e.g. 2026-06-08T19:30:14).
    if "T" in text:
        date_part = text.split("T", 1)[0]
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(date_part, fmt).date()
            except ValueError:
                continue

    candidates = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y%m%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y.%m.%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y.%m.%d %H:%M",
        "%Y/%m/%d %H:%M",
    )
    for fmt in candidates:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    # Strip currency symbols, commas, whitespace.
    text = text.replace(",", "").replace("₩", "").replace("원", "").strip()
    if text in {"-", ""}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _find_header_row(rows: Sequence[Sequence[object]]) -> tuple[int, dict[str, int]]:
    """Locate the header row + column index map.

    Returns (row_index, column_map) where column_map has keys:
      "date", "description", "amount" (if present),
      "withdrawal", "deposit" (if present)
    Raises BankStatementParseError if no recognizable header found.
    """
    for r_idx, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS]):
        column_map: dict[str, int] = {}
        for c_idx, raw in enumerate(row):
            cell = _normalize_header_cell(raw)
            if not cell:
                continue
            matched_withdrawal = _header_matches(cell, _WITHDRAWAL_KEYWORDS)
            matched_deposit = _header_matches(cell, _DEPOSIT_KEYWORDS)
            if "date" not in column_map and _header_matches(cell, _DATE_KEYWORDS):
                column_map["date"] = c_idx
            if "description" not in column_map and _header_matches(cell, _DESC_KEYWORDS):
                column_map["description"] = c_idx
            if "memo" not in column_map and _header_matches(cell, _MEMO_KEYWORDS):
                column_map["memo"] = c_idx
            if "type" not in column_map and _header_matches(cell, _TYPE_KEYWORDS):
                column_map["type"] = c_idx
            if "withdrawal" not in column_map and matched_withdrawal:
                column_map["withdrawal"] = c_idx
            if "deposit" not in column_map and matched_deposit:
                column_map["deposit"] = c_idx
            # "출금액"/"입금액" also contain "금액", so they'd match the generic
            # amount keyword too. Never classify a withdrawal/deposit column as
            # the signed `amount` column — otherwise a statement with separate
            # 출금액/입금액 columns reads only one of them and drops the other
            # (e.g. deposit-only 학생회비 rows → 0 parsed → empty reconciliation).
            if (
                "amount" not in column_map
                and not matched_withdrawal
                and not matched_deposit
                and _header_matches(cell, _AMOUNT_KEYWORDS)
            ):
                column_map["amount"] = c_idx
        has_amount_signal = (
            "amount" in column_map
            or "withdrawal" in column_map
            or "deposit" in column_map
        )
        if "date" in column_map and has_amount_signal:
            # Description is optional but common — fall back to empty string at parse time.
            return r_idx, column_map
    raise BankStatementParseError(
        "거래내역 엑셀에서 헤더(거래일/금액 등)를 찾을 수 없습니다."
    )


def _column_index_from_cell_ref(ref: str) -> int:
    """Convert an Excel cell ref (e.g. 'C12') to a 0-based column index."""
    letters = "".join(ch for ch in ref if ch.isalpha())
    index = 0
    for ch in letters.upper():
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index - 1


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for si in root.iter():
        if not si.tag.endswith("}si") and si.tag != "si":
            continue
        parts: list[str] = []
        for node in si.iter():
            if node.tag.endswith("}t") or node.tag == "t":
                if node.text:
                    parts.append(node.text)
        strings.append("".join(parts))
    return strings


def _xlsx_cell_value(cell: ET.Element, shared_strings: Sequence[str]) -> object:
    cell_type = cell.attrib.get("t")
    value_node = None
    for node in cell:
        if node.tag.endswith("}v") or node.tag == "v":
            value_node = node
            break
    if value_node is None or value_node.text is None:
        return None
    raw = value_node.text
    if cell_type == "s":
        return shared_strings[int(raw)]
    if cell_type == "d":
        return raw
    if cell_type == "inlineStr":
        for node in cell:
            if node.tag.endswith("}t") or node.tag == "t":
                return node.text
        return None
    try:
        if "." in raw:
            return float(raw)
        return int(raw)
    except ValueError:
        return raw


def _read_rows_xlsx_via_xml(file_bytes: bytes) -> list[list[object]]:
    """Fallback reader for strict-OOXML exports (e.g. Kakao Bank email xlsx).

    Some bank exports use the `purl.oclc.org` namespace that openpyxl does not
    load as worksheets, leaving `wb.sheetnames` empty even though sheet1.xml
    contains the transaction table.
    """
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
            sheet_name = next(
                (name for name in zf.namelist() if name.startswith("xl/worksheets/sheet")),
                None,
            )
            if sheet_name is None:
                return []
            shared_strings = _xlsx_shared_strings(zf)
            root = ET.fromstring(zf.read(sheet_name))
    except (KeyError, OSError, ET.ParseError, zipfile.BadZipFile, ValueError) as exc:
        raise BankStatementParseError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    sparse_rows: dict[int, dict[int, object]] = {}
    max_col = 0
    for row_el in root.iter():
        if not (row_el.tag.endswith("}row") or row_el.tag == "row"):
            continue
        row_idx = int(row_el.attrib.get("r", "0")) - 1
        if row_idx < 0:
            continue
        for cell in row_el:
            if not (cell.tag.endswith("}c") or cell.tag == "c"):
                continue
            ref = cell.attrib.get("r", "")
            if not ref:
                continue
            col_idx = _column_index_from_cell_ref(ref)
            sparse_rows.setdefault(row_idx, {})[col_idx] = _xlsx_cell_value(
                cell, shared_strings
            )
            max_col = max(max_col, col_idx)

    if not sparse_rows:
        return []

    rows: list[list[object]] = []
    for row_idx in sorted(sparse_rows):
        row_map = sparse_rows[row_idx]
        rows.append([row_map.get(col) for col in range(max_col + 1)])
    return rows


def _read_rows_xlsx(file_bytes: bytes) -> list[list[object]]:
    rows: list[list[object]] = []
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except (InvalidFileException, KeyError, OSError) as exc:
        raise BankStatementParseError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc
    try:
        if wb.sheetnames:
            sheet = wb[wb.sheetnames[0]]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not rows:
        rows = _read_rows_xlsx_via_xml(file_bytes)
    if not rows:
        raise BankStatementParseError("엑셀에 시트가 없습니다.")
    return rows


def _read_rows_xls(file_bytes: bytes) -> list[list[object]]:
    """Read a legacy binary .xls (BIFF) via xlrd; openpyxl cannot open these."""
    try:
        book = xlrd.open_workbook(file_contents=file_bytes)
    except Exception as exc:  # noqa: BLE001 — xlrd raises many error types on bad files
        raise BankStatementParseError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc
    if book.nsheets == 0:
        raise BankStatementParseError("엑셀에 시트가 없습니다.")
    sheet = book.sheet_by_index(0)
    rows: list[list[object]] = []
    for r in range(sheet.nrows):
        row: list[object] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            # Date-typed cells come back as Excel serials; hand _coerce_date a
            # real datetime. Everything else (text/number) passes through as-is.
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def parse_bank_statement_bytes(file_bytes: bytes) -> list[ParsedTransaction]:
    """Parse an xlsx/xls workbook into a ParsedTransaction list.

    Pure function for testing. Picks the reader by sniffing the file's magic
    bytes (OLE2 → legacy .xls via xlrd, otherwise .xlsx/.xlsm via openpyxl).
    """
    if file_bytes[:8] == _OLE2_MAGIC:
        rows = _read_rows_xls(file_bytes)
    else:
        rows = _read_rows_xlsx(file_bytes)
    if not rows:
        raise BankStatementParseError("엑셀이 비어 있습니다.")

    header_idx, columns = _find_header_row(rows)
    transactions: list[ParsedTransaction] = []

    for row in rows[header_idx + 1 :]:
        if not row or all(cell is None or cell == "" for cell in row):
            continue
        d = _coerce_date(_safe_get(row, columns.get("date")))
        if d is None:
            continue

        description_raw = _safe_get(row, columns.get("description"))
        description = str(description_raw).strip() if description_raw is not None else ""
        memo_raw = _safe_get(row, columns.get("memo"))
        memo = str(memo_raw).strip() if memo_raw is not None else ""
        if memo:
            description = f"{description} ({memo})" if description else memo

        amount = _resolve_amount(row, columns)
        if amount is None:
            continue

        transactions.append(
            ParsedTransaction(
                transaction_date=d,
                description=description[:255],
                amount=amount,
            )
        )

    return transactions


def _safe_get(row: Sequence[object], idx: int | None) -> object:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _resolve_amount(row: Sequence[object], columns: dict[str, int]) -> Decimal | None:
    """Resolve a single signed Decimal from either:
      - a Kakao-style `구분` + `거래금액` pair (출금 → negative),
      - a single signed `amount` column (used as-is), or
      - a withdrawal/deposit pair (withdrawal becomes negative).
    """
    if "type" in columns and "amount" in columns:
        type_text = str(_safe_get(row, columns["type"]) or "").strip()
        amount = _coerce_decimal(_safe_get(row, columns["amount"]))
        if amount is None:
            return None
        if "출금" in type_text:
            return -abs(amount)
        if "입금" in type_text:
            return abs(amount)
        return amount

    if "amount" in columns:
        return _coerce_decimal(_safe_get(row, columns["amount"]))

    withdrawal = _coerce_decimal(_safe_get(row, columns.get("withdrawal")))
    deposit = _coerce_decimal(_safe_get(row, columns.get("deposit")))

    if withdrawal and withdrawal != Decimal(0):
        return -abs(withdrawal)
    if deposit and deposit != Decimal(0):
        return abs(deposit)
    return None


async def create_upload_with_transactions(
    session: AsyncSession,
    *,
    settlement_id: uuid.UUID,
    stored_file: StoredBankStatement,
    transactions: Sequence[ParsedTransaction],
) -> BankStatementUpload:
    upload = BankStatementUpload(
        settlement_id=settlement_id,
        source_file_name=stored_file.original_name,
        source_file_path=str(stored_file.absolute_path),
        status=(
            BankStatementStatus.COMPLETED if transactions else BankStatementStatus.FAILED
        ),
        parsed_rows_count=len(transactions),
    )
    session.add(upload)
    await session.flush()

    for tx in transactions:
        session.add(
            BankTransaction(
                upload_id=upload.id,
                transaction_date=tx.transaction_date,
                description=tx.description,
                amount=tx.amount,
            )
        )

    await session.commit()
    await session.refresh(upload)
    return upload


async def get_upload_or_raise(
    session: AsyncSession, upload_id: uuid.UUID
) -> BankStatementUpload:
    upload = await session.get(BankStatementUpload, upload_id)
    if upload is None:
        raise BankStatementUploadNotFound("거래내역 업로드를 찾을 수 없습니다.")
    return upload


async def delete_bank_statement_upload(
    session: AsyncSession,
    *,
    upload: BankStatementUpload,
) -> None:
    """Hard-delete one upload, its parsed rows, linked reconciliation rows, and file."""
    tx_ids = list(
        await session.scalars(
            select(BankTransaction.id).where(BankTransaction.upload_id == upload.id)
        )
    )
    if tx_ids:
        await session.execute(
            delete(ReconciliationResult).where(
                ReconciliationResult.bank_transaction_id.in_(tx_ids)
            )
        )
        await session.execute(
            delete(BankTransaction).where(BankTransaction.upload_id == upload.id)
        )

    file_path = Path(upload.source_file_path)
    await session.delete(upload)
    await session.commit()
    file_path.unlink(missing_ok=True)


async def list_settlement_uploads(
    session: AsyncSession,
    *,
    settlement_id: uuid.UUID,
) -> Sequence[BankStatementUpload]:
    stmt = (
        select(BankStatementUpload)
        .where(BankStatementUpload.settlement_id == settlement_id)
        .order_by(BankStatementUpload.created_at.desc())
    )
    result = await session.scalars(stmt)
    return list(result.all())


async def list_settlement_transactions(
    session: AsyncSession,
    *,
    settlement_id: uuid.UUID,
) -> Sequence[BankTransaction]:
    """All bank transactions across every upload for the settlement."""
    stmt = (
        select(BankTransaction)
        .join(BankStatementUpload, BankTransaction.upload_id == BankStatementUpload.id)
        .where(BankStatementUpload.settlement_id == settlement_id)
        .order_by(BankTransaction.transaction_date.desc(), BankTransaction.created_at.desc())
    )
    result = await session.scalars(stmt)
    return list(result.all())
